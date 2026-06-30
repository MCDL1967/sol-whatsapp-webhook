"""
logs_export.py
Exports cases from the LOGS v2 SQLite DB into the live Excel workbooks.

Doctrine:
- DB is the source of truth
- Excel is a generated output / reporting artifact
- Before overwriting, a timestamped snapshot is saved alongside the live file
- Column order and sheet names match the existing live files

Usage:
    python logs_export.py gm
    python logs_export.py fnb
    python logs_export.py sec
    python logs_export.py adm
    python logs_export.py all
"""

import argparse
import shutil
from pathlib import Path
from datetime import datetime

import pytz
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from logs_db import get_db, init_db
from logs_models import PANAMA_TZ

# ── File paths ────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent

# ── Live workbook target paths ─────────────────────────────────────────────────
# These point at the real Google Drive mounted paths on the Mac.
# Override any individual path with an environment variable if needed:
#   LOGS_PATH_GM, LOGS_PATH_FNB, LOGS_PATH_SEC, LOGS_PATH_ADM
#
# Real Drive targets (Mac, MCDL1):
_GDRIVE = Path("/Users/MCDL1/Library/CloudStorage/GoogleDrive-sol.concierge.ai@gmail.com/My Drive/SOL_DEMO_1/LOGS")

_DEFAULT_PATHS = {
    "gm":  _GDRIVE / "GM"  / "gm_consolidated_log_live.xlsx",
    "fnb": _GDRIVE / "FNB" / "fnb_team_log_live.xlsx",
    "sec": _GDRIVE / "SEC" / "sec_team_log_live.xlsx",
    "adm": _GDRIVE / "ADM" / "adm_team_log_live.xlsx",
}

import os as _os
EXCEL_PATHS = {
    "gm":  Path(_os.environ.get("LOGS_PATH_GM",  str(_DEFAULT_PATHS["gm"]))),
    "fnb": Path(_os.environ.get("LOGS_PATH_FNB", str(_DEFAULT_PATHS["fnb"]))),
    "sec": Path(_os.environ.get("LOGS_PATH_SEC", str(_DEFAULT_PATHS["sec"]))),
    "adm": Path(_os.environ.get("LOGS_PATH_ADM", str(_DEFAULT_PATHS["adm"]))),
}

SNAPSHOT_DIR = BASE_DIR.parent / "logs_v2_runtime" / "export_snapshots"

# ── Column definitions per view ───────────────────────────────────────────────
# These define the column order written to each workbook.
# Adjust to match your actual live file column order if it differs.

GM_COLUMNS = [
    "record_id", "timestamp", "category", "section_team", "dept_target",
    "request_type", "workflow_stage", "status", "urgency", "severity",
    "guest_name", "contact_phone", "contact_email", "preferred_contact_method",
    "location_section", "venue_or_department",
    "service_date", "service_time", "party_size",
    "incident_type", "summary", "special_requests_or_notes",
    "assigned_manager_or_queue", "follow_up_required", "follow_up_type",
    "broadcast_channel", "broadcast_status", "client_message_status",
    "internal_resolution_status", "customer_confirmation_status", "closure_status",
    "resolution_summary", "resolved_by", "resolution_timestamp",
    "issue_reopened", "final_closure_timestamp",
    "survey_sent_timestamp", "survey_response_timestamp", "customer_feedback_score",
    "routing_reason", "complaint_category", "request_category",
    "operator_notes",
]

FNB_COLUMNS = [
    "record_id", "timestamp", "status", "urgency",
    "guest_name", "contact_phone", "contact_email", "preferred_contact_method",
    "venue_or_department", "service_date", "service_time", "party_size",
    "summary", "special_requests_or_notes",
    "assigned_manager_or_queue", "follow_up_required", "follow_up_type",
    "internal_resolution_status", "customer_confirmation_status", "closure_status",
    "resolution_summary", "resolved_by", "resolution_timestamp",
    "issue_reopened", "operator_notes",
]

SEC_COLUMNS = [
    "record_id", "timestamp", "status", "urgency", "severity",
    "guest_name", "contact_phone", "location_section",
    "incident_type", "summary", "special_requests_or_notes",
    "assigned_manager_or_queue",
    "internal_resolution_status", "closure_status",
    "resolution_summary", "resolved_by", "resolution_timestamp",
    "issue_reopened", "final_closure_timestamp", "operator_notes",
]

ADM_COLUMNS = [
    "record_id", "timestamp", "request_type", "status", "urgency",
    "guest_name", "contact_phone", "contact_email", "preferred_contact_method",
    "venue_or_department", "service_date", "service_time", "party_size",
    "summary", "special_requests_or_notes",
    "assigned_manager_or_queue", "follow_up_required",
    "internal_resolution_status", "customer_confirmation_status", "closure_status",
    "resolution_summary", "resolved_by", "resolution_timestamp",
    "issue_reopened", "operator_notes",
]

VIEW_COLUMNS = {
    "gm":  GM_COLUMNS,
    "fnb": FNB_COLUMNS,
    "sec": SEC_COLUMNS,
    "adm": ADM_COLUMNS,
}

# ── Query per view ─────────────────────────────────────────────────────────────
VIEW_QUERIES = {
    "gm": "SELECT * FROM cases ORDER BY timestamp DESC",
    "fnb": "SELECT * FROM cases WHERE dept_target = 'FNB' ORDER BY timestamp DESC",
    "sec": "SELECT * FROM cases WHERE dept_target = 'SEC' ORDER BY timestamp DESC",
    "adm": "SELECT * FROM cases WHERE dept_target = 'ADM' ORDER BY timestamp DESC",
}


# ── Styling helpers ────────────────────────────────────────────────────────────
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL  = PatternFill("solid", start_color="1A2B4A")   # dark navy
ALT_FILL     = PatternFill("solid", start_color="F2F4F8")
BORDER_SIDE  = Side(style="thin", color="CCCCCC")
CELL_BORDER  = Border(bottom=BORDER_SIDE)

STATUS_COLORS = {
    "open":                          "FFF9C4",  # soft yellow
    "in_progress":                   "BBDEFB",  # light blue
    "pending_customer_confirmation": "FFE0B2",  # light orange
    "closed":                        "C8E6C9",  # light green
    "reopened":                      "F8BBD0",  # light pink
}


def _make_snapshot(target_path: Path):
    """Copy the live file to the snapshot directory with a timestamp suffix."""
    if not target_path.exists():
        return
    SNAPSHOT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(PANAMA_TZ).strftime("%Y%m%d_%H%M%S")
    snap_name = f"{target_path.stem}_{ts}{target_path.suffix}"
    snap_path = SNAPSHOT_DIR / snap_name
    shutil.copy2(target_path, snap_path)
    print(f"  [snapshot] {snap_path.name}")


def _write_sheet(ws, rows: list[dict], columns: list[str]):
    """Write header + data rows to an openpyxl worksheet."""
    # Header row
    ws.append(columns)
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        is_alt = (row_idx % 2 == 0)
        status = (row.get("status") or "").lower()
        status_color = STATUS_COLORS.get(status)

        for col_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=False)

            if status_color:
                cell.fill = PatternFill("solid", start_color=status_color)
            elif is_alt:
                cell.fill = ALT_FILL

            cell.border = CELL_BORDER


def export_view(view: str) -> dict:
    """Export one view (gm | fnb | sec | adm) to its live Excel file."""
    target_path = EXCEL_PATHS[view]
    columns = VIEW_COLUMNS[view]
    query = VIEW_QUERIES[view]

    # Fetch from DB
    conn = get_db()
    rows = [dict(r) for r in conn.execute(query).fetchall()]
    conn.close()

    print(f"\n[export:{view.upper()}] {len(rows)} cases → {target_path.name}")

    # Snapshot before overwrite
    _make_snapshot(target_path)

    # Ensure parent dir exists
    target_path.parent.mkdir(parents=True, exist_ok=True)

    # Load existing workbook to preserve other sheets/formatting, or create new
    DATA_SHEET_NAME = "GM_Consolidated_Log"
    if target_path.exists():
        wb = load_workbook(str(target_path))
        # Use GM_Consolidated_Log if it exists, otherwise use first sheet
        if DATA_SHEET_NAME in wb.sheetnames:
            sheet_name = DATA_SHEET_NAME
        else:
            sheet_name = wb.sheetnames[0]
        del wb[sheet_name]
        ws = wb.create_sheet(sheet_name, wb.sheetnames.index(sheet_name) if sheet_name in wb.sheetnames else 0)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = DATA_SHEET_NAME

    _write_sheet(ws, rows, columns)

    # Add export metadata to a hidden sheet
    meta_sheet_name = "_export_meta"
    if meta_sheet_name in wb.sheetnames:
        del wb[meta_sheet_name]
    meta = wb.create_sheet(meta_sheet_name)
    meta.sheet_state = "hidden"
    now_str = datetime.now(PANAMA_TZ).strftime("%Y-%m-%dT%H:%M:%S")
    meta["A1"] = "exported_at"
    meta["B1"] = now_str
    meta["A2"] = "source"
    meta["B2"] = "LOGS_v2_SQLite"
    meta["A3"] = "view"
    meta["B3"] = view.upper()
    meta["A4"] = "row_count"
    meta["B4"] = len(rows)

    wb.save(str(target_path))
    print(f"  [export:{view.upper()}] Saved: {target_path}")

    return {"view": view, "rows": len(rows), "path": str(target_path)}


def main():
    parser = argparse.ArgumentParser(description="Export LOGS v2 DB to Excel workbooks")
    parser.add_argument("target", choices=["gm", "fnb", "sec", "adm", "all"],
                        help="Which view to export")
    args = parser.parse_args()

    init_db()

    targets = ["gm", "fnb", "sec", "adm"] if args.target == "all" else [args.target]

    results = []
    for view in targets:
        result = export_view(view)
        results.append(result)

    print(f"\n[logs_export] Complete.")
    for r in results:
        print(f"  {r['view'].upper()}: {r['rows']} rows → {r['path']}")


if __name__ == "__main__":
    main()
